from collections import Counter, defaultdict

from analysis.models import AssessmentMethod
from annotations.constants import SAF_FASES_COLUMN, SAF_UTT_HEADER
from natsort import natsorted
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from sastadev.allresults import AllResults
from sastadev.reduceresults import exact2results

QUERYCOUNT_HEADERS = ['Query', 'Item',
                      SAF_FASES_COLUMN, SAF_UTT_HEADER, 'Matches']

TOTAL_LABEL = 'totaal'
NOT_APPLICABLE_LABEL = 'nvt'


def querycounts_to_xlsx(allresults: AllResults, method: AssessmentMethod):
    wb = Workbook()
    worksheet = wb.active

    # header
    worksheet.append(QUERYCOUNT_HEADERS)
    header = worksheet["1:1"]
    for cell in header:
        cell.font = Font(bold=True)

    nonempty_queries = {k: v for k, v in allresults.exactresults.items() if v}

    res = exact2results(nonempty_queries)

    # need to reduce the results
    # because of double results for lemma queries
    reduced_results = defaultdict(Counter)
    for (k, _), v in res.items():
        reduced_results[k] += v

    # write rows of data
    for qid in natsorted(reduced_results):
        # get query info
        cntr = reduced_results[qid]
        q = method.queries.get(query_id=qid)

        # write the total row
        total_row = [qid, q.item, q.fase or NOT_APPLICABLE_LABEL,
                     TOTAL_LABEL, sum(cntr.values())]
        worksheet.append(total_row)

        for utt_id in natsorted(cntr.keys()):
            row = [None, None, None, utt_id, cntr[utt_id]]
            worksheet.append(row)

    worksheet.auto_filter.ref = worksheet.dimensions

    # column widths
    autosize_columns(worksheet)

    return wb


def autosize_columns(worksheet) -> None:
    dim_holder = DimensionHolder(worksheet=worksheet)
    for col in range(worksheet.min_column, worksheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            worksheet, min=col, max=col, auto_size=True)
    worksheet.column_dimensions = dim_holder
