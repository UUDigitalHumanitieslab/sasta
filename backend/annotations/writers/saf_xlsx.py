import itertools
from dataclasses import dataclass, field
from io import BytesIO
from typing import Dict, List, Tuple

from analysis.models import MethodCategory
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from sastadev.allresults import AllResults, ResultsKey
from sastadev.methods import Method
from sastadev.sastatypes import ExactResults
from annotations.constants import (POST_WORDS_HEADERS, PRE_WORDS_HEADERS,
                                   SAF_COMMENT_LEVEL, SAF_UTT_LEVEL)
from annotations.utils import autosize_columns, format_worksheet, get_max_words, ljust


@dataclass
class SAFWriter():
    method: Method
    results: AllResults
    workbook: Workbook = field(init=False, default_factory=Workbook)
    anno_ws: Worksheet = field(init=False)
    method_category: MethodCategory = field(init=False)
    # Number of non-word columns, counted from 0
    word_offset: int = field(default=len(PRE_WORDS_HEADERS), init=False)
    # Number of words of the longest utterance in the results
    max_words: int = field(init=False)
    # Offset for row (1 - len(levels)) below utt row (0)
    level_offsets: Dict[str, int] = field(init=False)
    # Number of rows each utterance takes up (utt + level rows)
    utt_n_rows: int = field(init=False)
    # header row
    anno_headers: List[str] = field(init=False)

    def __post_init__(self) -> None:
        self.max_words = get_max_words(self.results)
        self.method_category = MethodCategory.objects.get(
            name=self.method.name.upper())

        all_levels = [SAF_UTT_LEVEL,
                      *self.method_category.levels,
                      SAF_COMMENT_LEVEL]
        self.level_offsets = {
            level.lower(): index
            for (index, level)
            in enumerate(all_levels)
        }
        self.utt_n_rows = (len(all_levels))
        self.anno_headers = self._annotations_header_row()
        self.make_workbook()

    def write(self, target: BytesIO) -> None:
        '''Write the completed output file'''
        self.workbook.save(target)

    def make_workbook(self) -> None:
        '''Create the complete workbook.
        Any additional required sheets should created in this method.
        '''
        _ = self._make_annotations_worksheet()
        format_worksheet(self.anno_ws)
        autosize_columns(self.anno_ws)
        errors_ws = self.workbook.create_sheet('error', 1)
        errors_ws.cell(1, 1).value = 'Hier komen errors'

    def _make_annotations_worksheet(self) -> Worksheet:
        '''Transform results into a SAF-formatted worksheet'''
        self.anno_ws = self.workbook.create_sheet(title='annotations', index=0)
        # Headers
        self.anno_ws.append(self.anno_headers)

        # Rest
        self._make_levels_rows(self.anno_ws)

        # Fill with values
        for qid, qresults in self.results.exactresults.items():
            self._fill_query(qid, qresults)
        return self.anno_ws

    def _annotations_header_row(self) -> List[str]:
        '''Create header row with correct number of word columns'''
        word_headers = [f'Word{i}' for i in range(1, self.max_words + 1)]

        return list(itertools.chain(
            PRE_WORDS_HEADERS,
            word_headers,
            POST_WORDS_HEADERS
        ))

    def _uttlevel_row(self, id: int, words: List[str]) -> List[str]:
        '''Create utterance level row'''
        pre_word_values = [id, SAF_UTT_LEVEL, None]  # Unaligned
        word_values = ljust(words, self.max_words)
        post_word_values = [None, None]  # Fases, Commentaar
        return list(itertools.chain(
            pre_word_values,
            word_values,
            post_word_values
        ))

    def _make_levels_rows(self, ws: Worksheet) -> None:
        '''Create rows for all utterances, all levels'''
        row_size = len(self.anno_headers)
        all_levels = self.method_category.levels + [SAF_COMMENT_LEVEL]

        for utt_id, words in sorted(self.results.allutts.items(),
                                    key=lambda x: x[0]):
            ws.append(self._uttlevel_row(utt_id, words))
            for level in all_levels:
                level_row = ljust([utt_id, level], row_size)
                ws.append(level_row)

    def _fill_query(self, query_id: ResultsKey, exact_results: ExactResults):
        '''Find and fill all cells for a single query'''
        lemma_item = None
        if isinstance(query_id, Tuple) and not query_id[0] == query_id[1]:
            # Lemma queries hold the lemma in second position
            lemma_item = query_id[1]

        simple_query_id = query_id[0]
        query = self.method.queries.get(simple_query_id)
        item = lemma_item or query.item
        fase = query.fase

        for utt_id, word_nr in exact_results:
            # We cannot assume that utterances are numbered 1-N sequentially
            try:
                utt_nr = list(self.results.allutts.keys()).index(utt_id)
            except ValueError:
                utt_nr = list(self.results.allutts.keys()).index(int(utt_id))
            row, col = self._cell_location(utt_nr, query.level, word_nr)
            cell = self.anno_ws.cell(row, col)
            self._append_item(cell, item)
            if fase:
                self._append_fase(row, str(fase))

    def _cell_location(self, utt_nr: int, level: str,
                       word_nr: int) -> Tuple[int, int]:
        '''Find the coordinates of a cell'''
        return (
            self._uttlevel_row_number(utt_nr, level),
            self._word_col_number(word_nr)
        )

    def _uttlevel_row_number(self, utt_nr: int, level: str) -> int:
        '''Calculate the row number for level of utterance (1 indexed)'''
        total = 1  # header and 1-indexed offsets
        utt_offset = (utt_nr * self.utt_n_rows) + 1
        level_offset = self.level_offsets.get(level.lower(), 0)
        total += utt_offset + level_offset
        return total

    def _word_col_number(self, word_nr: int) -> int:
        '''Calculate the column number for a word'''
        return word_nr + len(PRE_WORDS_HEADERS)

    def _append_item(self, cell: Cell, item: str) -> None:
        cell.value = item if not cell.value else f'{cell.value}, {item}'

    def _append_fase(self, row: int, fase: str) -> None:
        fase_cell = self.anno_ws.cell(row, len(self.anno_headers) - 1)
        sep = ', '
        if not fase_cell.value:
            fase_cell.value = fase
        else:
            current = set(fase_cell.value.split(sep))
            current.add(fase)
            new = sep.join(sorted(list(current)))
            fase_cell.value = new
