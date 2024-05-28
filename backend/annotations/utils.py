from typing import Any, List, Optional

from annotations import constants
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.protection import Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from sastadev.allresults import AllResults


def preflabel(labels: List[str], casing: Optional[callable] = None) -> str:
    try:
        label = labels[0]
        return casing(label) if casing else label
    except IndexError:
        return ''


def get_max_words(allresults: AllResults) -> int:
    '''Get the length of the longest utterance in the results'''
    return max(len(v) for v in allresults.allutts.values())


def ljust(li: List[Any], n: int, fillvalue=None) -> List[Any]:
    '''Pad the list with fillvalues up to N'''
    return li + [fillvalue] * (n - len(li))


def format_worksheet(worksheet) -> None:
    '''Locks all cells except annotation fields.
    Gives utterance rows a primary background.
    '''

    # start by locking the entire sheet
    worksheet.protection.sheet = True
    unlocked = Protection(locked=False)

    header = worksheet["1:1"]
    for cell in header:
        # bold headers
        cell.font = Font(bold=True)

    # yelow background for each utterance row
    for row in list(worksheet.rows)[1:]:
        if row[1].value == constants.SAF_UTT_LEVEL:
            for cell in row:
                cell.font = Font(color='FFFFFF')
                cell.fill = PatternFill(
                    start_color=constants.PRIMARY_COLOR,
                    end_color=constants.PRIMARY_COLOR,
                    fill_type="solid")
        else:
            # unlock non-utterance rows
            # skip the first two columns (utt number and level)
            for cell in row[2:]:
                cell.protection = unlocked


def autosize_columns(worksheet) -> None:
    dim_holder = DimensionHolder(worksheet=worksheet)
    for col in range(worksheet.min_column, worksheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            worksheet, min=col, max=col, auto_size=True)
    worksheet.column_dimensions = dim_holder


def cast_to_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    elif isinstance(value, str):
        return value == 'yes'
