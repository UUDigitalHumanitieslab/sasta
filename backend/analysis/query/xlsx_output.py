from collections import Counter
from typing import List, Tuple

from analysis.models import AssessmentMethod
from analysis.results.results import AllResults
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

QUERYCOUNT_HEADERS = ['Query', 'Item', 'Phase', 'Utterance', 'Matches']


def querycounts_to_xlsx(allresults: AllResults, method: AssessmentMethod):
    all_data = dict(allresults.coreresults, **allresults.postresults)

    wb = Workbook()
    worksheet = wb.active

    # header
    worksheet.append(QUERYCOUNT_HEADERS)
    header = worksheet["1:1"]
    for cell in header:
        cell.font = Font(bold=True)

    query_mapping = {
        q.query.id: (q.query.fase or 0, q.query.item)
        for q in queries
        if q.query.id in all_data
    }
    sorted_queries = sorted(
        sorted(
            query_mapping.items(),
            key=lambda item: item[0]
        ),
        key=lambda item: item[1][0]
    )

    for qid, (fase, item) in sorted_queries:
        fase = fase if fase else 'nvt'
        data = all_data[qid]

        if isinstance(data, int):
            row = [qid, item, fase, 'total', data]
            worksheet.append(row)
        elif isinstance(data, Counter):
            first_row = [qid, item, fase, 'total', sum(data.values())]
            worksheet.append(first_row)
            for utt in sorted(data):
                if isinstance(utt, Tuple):
                    row = [None, None, None, utt[-1], data[utt]]
                else:
                    row = [None, None, None, utt, data[utt]]
                worksheet.append(row)

    worksheet.auto_filter.ref = worksheet.dimensions

    # column widths
    autosize_columns(worksheet)

    return wb

def autosize_columns(worksheet) -> None:
    dim_holder = DimensionHolder(worksheet=worksheet)
    for col in range(worksheet.min_column, worksheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            worksheet, min=col, max=col, auto_size=True)
    worksheet.column_dimensions = dim_holder
